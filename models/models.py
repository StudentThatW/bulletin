# -*- coding: utf-8 -*-

from odoo import models, fields, api
# работа с Excel
from openpyxl import load_workbook 
from lxml import etree
from odoo.exceptions import ValidationError
from datetime import datetime, timedelta

import base64
import psycopg2
import string
import csv 
import os
import re

# модель Вид статистики
class bulletin_industries_economy(models.Model):
    _name = 'bulletin.industries_economy'

    name = fields.Char( string = "Отрасли" )
    industries_stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    industries_parameter = fields.Many2one( 'bulletin.industries_parameter', string = 'Параметры статистики' ) 
    datalens = fields.Char( string = "Таблица в DataLens" )  
    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )

# модель Регионы regions
class bulletin_regions(models.Model):
    _name = 'bulletin.regions'

    reg_number = fields.Integer( string = "№ рег." )                                    
    name = fields.Char( string = "Регион" )                                             
    archive = fields.Boolean( string = 'Архив' )                                        
    eng_name = fields.Char( string = "Росстат"  )                                       
    federal = fields.Many2one( 'bulletin.federal', string = 'Федеральный округ' )       
    fedstat = fields.Char( string = "Наименование на fedstat.ru" )                      
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона' )     
    reg_name = fields.Char( string = "Сокр. наим." )                                    
    is_ind = fields.Boolean( string = 'Индустриальный' )                                
    federal_number = fields.Integer( string = "№ рег в ФО" )   

# модель Федеральный округ
class bulletin_federal(models.Model):
    _name = 'bulletin.federal'

    fed_number = fields.Integer( string = "№ округа" )
    name = fields.Char( string = "Федеральный округ" )

    fed_number_str = fields.Char( string = "№ округа", compute = "get_fed_number_str"  )

    def get_fed_number_str(self):
        for record in self:
            if record.fed_number:
                record.fed_number_str = str( record.fed_number )

    fedstat = fields.Char( string = "Наименование на fedstat.ru" )

# модель Работник workers
class bulletin_workers(models.Model):
    _name = 'bulletin.workers'

    name = fields.Char( string = 'Сотрудник' )
    is_active = fields.Boolean( string = 'Активный' )
    user = fields.Many2one( 'res.users', string = 'Логин' )
    departament = fields.Many2one( 'bulletin.departament', string = 'Подразделение' )
    job = fields.Many2one( 'bulletin.job', string = 'Должность' )
    bulletin = fields.Many2one( 'bulletin.bulletin', string='bulletin' )

# модель Подразделения    
class bulletin_departament(models.Model):
    _name = 'bulletin.departament'

    id = fields.Integer( string = "id" )
    name = fields.Char( string='Подразделение' )

    manager = fields.Many2one( 'bulletin.workers', string = 'Руководитель' )
    parent  = fields.Many2one( 'bulletin.departament', string = 'Вышестоящее подразделение' )
    workers_ids = fields.One2many( 'bulletin.workers', 'departament', string = 'Сотрудники' )
    workers_count = fields.Integer(  string = "Кол-во сотрудников", compute = "get_workers_count" )

    def get_workers_count(self):
        for record in self:
            record.workers_count = self.env[ 'bulletin.workers' ].search_count( [ ( 'departament', '=', record.id) ] )

# модель Должности    
class bulletin_job(models.Model):
    _name = 'bulletin.job'
    name = fields.Char( string = 'Должность' )

# модель Показатели отрасли 
class bulletin_indicators_industries(models.Model):
    _name = 'bulletin.indicators_industries'

    name = fields.Char( string = "Показатели отрасли" )
    upper_name = fields.Char( string = "Показатели отрасли" )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Отрасль' )
    fedstat = fields.Char( string = "Наименование на fedstat.ru" )
    sheet_name = fields.Char( string = "Лист Excel" )
    rf = fields.Integer( string = "РФ * 1000" )  

# модель Характеристики отрасли 
class bulletin_characteristics_industries(models.Model):
    _name = 'bulletin.characteristics_industries'

    name = fields.Char( string = "Характеристики отрасли" )
    upper_name = fields.Char( string = "Характеристики отрасли" )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Отрасль' )
    fedstat = fields.Char( string = "Наименование на fedstat.ru" )

# модель Тип региона 
class bulletin_region_type(models.Model):
    _name = 'bulletin.region_type'

    name = fields.Char( string = "Тип региона" )

# модель тип статистики 
class bulletin_stat_type(models.Model):

    _name = 'bulletin.stat_type'
    _description = 'Тип статистики'

    name = fields.Char( string = "Тип статистики" )

# модель стат. периоды
class bulletin_load_date(models.Model):

    _name = 'bulletin.load_date'
    _description = 'Периоды'

    name = fields.Char( string = 'Период' )
    load_date = fields.Date( string = 'Дата периода' ) 
    comment = fields.Char( string = 'Примечание' )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Вид статистики' )
    quarter = fields.Boolean( string = 'Квартал' )
    bulletin_bulletin = fields.Many2one( 'bulletin.bulletin', string = 'Статистика' )
    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )

# модель ОКВЭД2 
class bulletin_okved(models.Model):
    
    _name = 'bulletin.okved'
    _description = 'ОКВЭД2'

    name = fields.Char( string = "Наименование" )
    code = fields.Char( string = "Код" ) 
    status = fields.Boolean( string = 'Активен' )
    short_name = fields.Char( string = "Наименование кратко" )

# модель parameter_type тип параметра 
class bulletin_parameter_type(models.Model):

    _name = 'bulletin.parameter_type'
    _description = 'Тип параметра'

    name = fields.Char( string = "Тип параметра" )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    sheet_name = fields.Char( string = "Лист Excel" )
    rf = fields.Integer( string = "РФ * 1000" )   

# модель Варианты наименования региона regions_name
class bulletin_regions_name(models.Model):

    _name = 'bulletin.regions_name'
    _description = 'Варианты наименования региона'

    name = fields.Char( string = "Вариант наименования" ) 
    region = fields.Many2one( 'bulletin.regions', string = 'Регион' ) 

# модель Периодичность статистики statistics_period
class bulletin_statistics_period(models.Model):

    _name = 'bulletin.statistics_period'
    _description = 'Периодичность статистики'

    name = fields.Char( string = "Периодичность статистики" )

# модель Параметры статистики industries_parameter
class bulletin_industries_parameter(models.Model):
    _name = 'bulletin.industries_parameter'
    _description = 'Параметры статистики'

    name = fields.Char( string = "Параметры статистики" )

# модель Статистика statistics
class bulletin_statistics(models.Model):

    _name = 'bulletin.statistics'
    _description = 'Статистика'

    _order = "date_stat, okved, count"        # desc

    name = fields.Char( string = "Статистика" )
    date_stat = fields.Date( string = "Дата" ) 
    quarter = fields.Boolean( string = 'Квартал' )
    count = fields.Float( digits=( 10, 3 ), string = "Количество" )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатели отрасли' )
    characteristics_industries = fields.Many2one( 'bulletin.characteristics_industries', string = 'Характеристики отрасли' )
    region = fields.Many2one( 'bulletin.regions', string = 'Регион' )  
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона' ) 
    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2' )
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )
    stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    column = fields.Char( string = "Колонка" )
    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )

# модель rating Рейтинг Виды статистик и параметры 
class bulletin_rating_industries(models.Model):

    _name = 'bulletin.rating_industries'
    _description = 'Рейтинг Виды статистик и параметры'

    name = fields.Char( string = "Наименование" )

    rating = fields.Many2one( 'bulletin.rating', string = 'Рейтинг' )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра', domain = "[('industries', '=', industries)]" )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатели отрасли', domain = "[('industries', '=', industries)]" )
    is_activ = fields.Boolean( string = 'Активно' )
    number_industries = fields.Integer( string = "№ пок" )

    date_start = fields.Date( string = "Дата начала" ) 
    date_end = fields.Date( string = "Дата окончания" ) 

    # По изменению поля industries установить отбор для полей parameter_type, indicators_industries - multi domain ! 
    @api.onchange( 'industries' )
    def change_industries( self ):

        # удалить старые значения в полях, .т.к. у новой измененной статистики этих параметров может не быть совсем
        self.indicators_industries = None
        self.parameter_type = None

        # установить отбор
        domain_start = { 'date_start_view': [] }  # пустой список 
        for rec in self:
            if rec.industries:
                rating_ctx = int( self.env.context.get( 'rating_ctx' ) )
                rec.rating = self.env[ 'bulletin.rating' ].search( [ ( 'id', '=', rating_ctx) ] ) 
                domain_start = { 'parameter_type': [ ('industries.id', '=', rec.industries.id ) ], 'indicators_industries': [ ('industries.id', '=', rec.industries.id ) ] }             
        return { 'domain': domain_start }
        # -- установить отбор

# модель rating Рейтинг 
class bulletin_rating(models.Model):

    _name = 'bulletin.rating'
    _description = 'Рейтинг'

    name = fields.Char( string = "Рейтинг" )
    datalens_table = fields.Char( string = "Таблица в DataLens" ) 
    # локальный сервер
    loval_server = fields.Boolean( string = 'Локальный сервер', default = True )

    # ссылка для table
    table_ids = fields.One2many( 'bulletin.rating_table', 'rating',  nolabel = "1" )
    # ссылка для Рейтинг Виды статистик и параметры
    rating_industries_ids = fields.One2many( 'bulletin.rating_industries', 'rating',  nolabel = "1" )

    # не используется
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    date_1 = fields.Date( string = "Дата 1" ) 
    date_2 = fields.Date( string = "Дата 2" ) 

    # рейтинг в DataLens
    def rating_to_result( self ):

        # connect к базе result
        conn = ConnectToBase( self.loval_server )
        cursor = conn.cursor()

        # предварительно удалить всё из рейтинга
        sql = ( "DELETE FROM " + self.datalens_table  )
        cursor.execute( sql )
        conn.commit()
        
        # данные текущего рейтинга self.id
        ComStr = "SELECT region, federal, industries, date, count, delta, rang, reg_number, reg_name, federal_number, number_industries FROM bulletin_rating_table WHERE rating = " + str( self.id )
        self._cr.execute( ComStr )

        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        industries = self.industries.name

        q = 0
        while q < CountQ:          

            region = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', dataQ[q][0] ) ] )
            federal = self.env[ 'bulletin.federal' ].search( [ ( 'id', '=', dataQ[q][1] ) ] )
            industries = self.env[ 'bulletin.industries_economy' ].search( [ ( 'id', '=', dataQ[q][2] ) ] )

            date  = dataQ[q][3]
            count = dataQ[q][4]
            delta = dataQ[q][5]
            rang  = dataQ[q][6]
            reg_number = dataQ[q][7]
            reg_name   = dataQ[q][8]
            federal_number    = dataQ[q][9]
            number_industries = dataQ[q][10]

            # запись в retail
            MassivData = [ region.name, federal.name, industries.name, date, count, delta, rang, reg_number, reg_name, federal_number, number_industries ]
            sql = "INSERT INTO " + self.datalens_table + " ( region, federal, industries, date, count, delta, rang, reg_number, reg_name, federal_number, number_industries ) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s )"
            cursor.execute( sql, MassivData )
            conn.commit()

            q = q + 1

        conn.close()

    # очистить текущий рейтинг
    def clear_rating( self ):

        rec = self.env[ 'bulletin.rating_table' ].search( [ ( 'rating', '=', self.id ) ] )
        rec.unlink()

    # найти max дату статистики
    def rating_perid( self ):
                
        ComStr = "SELECT industries, id, parameter_type, indicators_industries, number_industries FROM bulletin_rating_industries WHERE is_activ = True and rating = " + str( self.id ) 
        self._cr.execute( ComStr )

        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        q = 0
        while q < CountQ:

            industries_id = str( dataQ[q][0] )
            id_id = int( dataQ[q][1] )

            parameter_type_id = dataQ[q][2]
            indicators_industries_id = dataQ[q][3]

            ComStrMax = 'SELECT MAX( date_stat ) FROM bulletin_statistics WHERE industries = ' + industries_id

            if parameter_type_id != None:
                ComStrMax = ComStrMax + " and parameter_type = '" + str( parameter_type_id ) + "'" 
            if indicators_industries_id != None:
                ComStrMax = ComStrMax + " and indicators_industries = '" + str( indicators_industries_id ) + "'" 

            self._cr.execute( ComStrMax )
            dataMax = self._cr.fetchall()

            # записать max даты
            self.env[ 'bulletin.rating_industries' ].search( [ ( 'id', '=', id_id ) ] ).write( {'date_end' : dataMax[0][0] } )

            q = q + 1

    # сформировать рейтинг
    def write_to_table( self ):

        # из ТЧ Настройка список показателей ( статистик ) с установленными параметрами и датами

        IndStr = "SELECT industries, indicators_industries, parameter_type, number_industries, date_start, date_end FROM bulletin_rating_industries WHERE is_activ = True and rating = " + str( self.id ) 
        self._cr.execute( IndStr )

        IndCountQ = self._cr.rowcount
        InddataQ = self._cr.fetchall()

        Ind = 0        
        while Ind < IndCountQ:

            industries_id = int( InddataQ[Ind][0] )
            indicators_industries_id = InddataQ[Ind][1]
            parameter_type_id = InddataQ[Ind][2] 
            number_industries = InddataQ[Ind][3]

            date_start = InddataQ[Ind][4] 
            date_end = InddataQ[Ind][5]

            # первая часть рейтинга - на дату date_start
            # raise ValidationError( date_start) 

            ComStr = "SELECT reg.id, stat.count FROM bulletin_regions AS reg, bulletin_statistics AS stat WHERE reg.id = stat.region and reg.reg_number > 0" 

            ComStr = ComStr + " and stat.date_stat = '" + str( date_start ) + "'"  
            ComStr = ComStr + " and stat.industries = '" + str( industries_id ) + "'" 

            if indicators_industries_id != None:
                ComStr = ComStr + " and stat.indicators_industries = '" + str( indicators_industries_id ) + "'" 
            if parameter_type_id != None:
                ComStr = ComStr + " and stat.parameter_type = " + str( parameter_type_id )  

            ComStr = ComStr + " ORDER BY stat.count DESC" 
            self._cr.execute( ComStr )

            CountQ = self._cr.rowcount
            dataQ = self._cr.fetchall()

            q = 0
            Rang = 1
            while q < CountQ:

                region_id = int( dataQ[q][0] )
                count = int( dataQ[q][1] )
                region = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', region_id ) ] )

                record_value = { 'rating':self.id, 'region':region_id, 'date':date_start, 'industries':industries_id, 'count':count, 'federal':region.federal.id, 'reg_number':region.reg_number, 'reg_name':region.reg_name, 'rang':Rang, 'federal_number':region.federal_number, 'number_industries':number_industries }
                self.env[ 'bulletin.rating_table' ].create( record_value )

                q = q + 1
                Rang = Rang + 1

            # вторая часть рейтинга - на дату date_end

            ComStr = "SELECT reg.id, stat.count FROM bulletin_regions AS reg, bulletin_statistics AS stat WHERE reg.id = stat.region and reg.reg_number > 0" 

            ComStr = ComStr + " and stat.date_stat = '" + str( date_end ) + "'"  
            ComStr = ComStr + " and stat.industries = '" + str( industries_id ) + "'" 

            if indicators_industries_id != None:
                ComStr = ComStr + " and stat.indicators_industries = '" + str( indicators_industries_id ) + "'" 
            if parameter_type_id != None:
                ComStr = ComStr + " and stat.parameter_type = " + str( parameter_type_id )  

            ComStr = ComStr + " ORDER BY stat.count DESC" 
            self._cr.execute( ComStr )

            CountQ = self._cr.rowcount
            dataQ = self._cr.fetchall()

            q = 0
            Rang = 1
            while q < CountQ:

                region_id = int( dataQ[q][0] )
                count = int( dataQ[q][1] )
                region = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', region_id ) ] )

                RangOne = self.env[ 'bulletin.rating_table' ].search( [ ( 'rating', '=', self.id ), ( 'region', '=', region_id ), ( 'industries', '=', industries_id ), ( 'date', '=', date_start ) ] )
                Delta = RangOne.rang - Rang

                record_value = { 'rating':self.id, 'region':region_id, 'date':date_end, 'industries':industries_id, 'count':count, 'federal':region.federal.id, 'reg_number':region.reg_number, 'reg_name':region.reg_name, 'rang':Rang, 'delta':Delta, 'federal_number':region.federal_number, 'number_industries':number_industries  }
                self.env[ 'bulletin.rating_table' ].create( record_value )

                q = q + 1
                Rang = Rang + 1

            Ind = Ind + 1

# модель rating Рейтинг тч 
class bulletin_rating_table(models.Model):

    _name = 'bulletin.rating_table'
    _description = 'Рейтинг тч'

    rating = fields.Many2one( 'bulletin.rating', string = 'Рейтинг' )

    region = fields.Many2one( 'bulletin.regions', string = 'Регион' ) 
    federal = fields.Many2one( 'bulletin.federal', string = 'Федеральный округ' )  
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Показатель' )
    date = fields.Date( string = "Дата" ) 
    count = fields.Float( digits = ( 10, 3 ), string = "Значение" )
    delta = fields.Integer( string = "Дельта" )
    rang = fields.Integer( string='Ранги' )
    reg_number = fields.Integer( string = "№ рег." )                
    reg_name = fields.Char( string = "Сокр. наим." )                
    federal_number = fields.Integer( string = "№ рег в ФО" )
    number_industries = fields.Integer( string = "№ пок" )      

# модель ИПП ipp
class bulletin_ipp(models.Model):

    _name = 'bulletin.ipp'
    _description = 'ИПП'

    # _order = "date_stat, okved, count"        # desc

    name = fields.Char( string = "Статистика" )

    # период
    date_stat = fields.Date( string = "Дата" ) 

    count = fields.Float( digits=( 10, 3 ), string = "Количество" )

    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )

    region = fields.Many2one( 'bulletin.regions', string = 'Регион' )  
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона' ) 

    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2' )

    # stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )

    # column = fields.Char( string = "Колонка" )

# модель bulletin.bulletin 
class bulletin_bulletin(models.Model):

    _name = 'bulletin.bulletin'
    _description = 'Начальная страница'

    name = fields.Char( string = "Наименование" )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )

    date_start = fields.Date( string = "Дата начала" ) 
    date_finish = fields.Date( string = "Дата окончания" ) 
    count_record = fields.Float( digits=( 10, 3 ), string = "Количество записей" )

    id = fields.Integer( string = "id" )

    record_type = fields.Integer( string = "record_type" )
    user = fields.Many2one( 'bulletin.workers', string = 'Загружено пользователем' )
    user_id = fields.Integer( string = "user_id"  )

    @api.onchange( 'user' )
    def change_operations_date( self ):
        for rec in self:
            if rec.user:
                rec.user_id = self.user.user.id  

    comment = fields.Char( string = "Таблица в DataLens" ) 
    multicomment = fields.Text( string = "Комментарий" ) 

    stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    link = fields.Char( string = "Ссылка" ) 

    # ссылка для attachment
    attachment_ids = fields.One2many( 'bulletin.attachment', 'bulletin', string = 'Присоединенные файлы' )
    # ссылка для load_date
    industries_ids = fields.One2many( 'bulletin.load_date', 'bulletin_bulletin', string = 'Периоды статистики' )

    # открыть tree модели bulletin с записями текущего пользователя
    def open_record(self): 

        # текущий пользователь
        current_user_id = self.env.uid

        views = [ (self.env.ref('bulletin.diagram_list').id, 'tree'), (self.env.ref('bulletin.diagram_form').id, 'form')  ]
        tree_view = {
            'name': 'Статистика',
            'view_mode': 'tree',
            'views': views,
            'res_model': 'bulletin.diagram',
            'type': 'ir.actions.act_window',   
            'domain': [ ('user.user', '=', current_user_id ) ],      # , ('operations_date', '=', work_date )
        }
        return tree_view 

    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )

    # найти в bulletin.load_date периоды соответствующие текущей записи bulletin_bulletin и удалить
    def del_perid( self ):                
        rec = self.env[ 'bulletin.load_date' ].search( [ ( 'bulletin_bulletin', '=', self.id ) ] )
        rec.unlink()

    # сгруппировать статистику по датам 
    def group_perid( self ):

        rec = self.env[ 'bulletin.load_date' ].search( [ ( 'bulletin_bulletin', '=', self.id ) ] )
        rec.unlink()
                
        ComStr = "SELECT date_stat, period, quarter FROM bulletin_statistics WHERE industries = " + str( self.industries.id ) + " GROUP BY date_stat, period, quarter ORDER BY date_stat DESC"
        self._cr.execute( ComStr )

        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        q = 0
        while q < CountQ:
            # заполнить даты начала - конца
            if q == 0:
                self.date_finish =dataQ[q][0]
            if q == CountQ - 1:
                self.date_start = dataQ[q][0]

            date_stat = str( dataQ[q][0] )
            industries_id = self.industries.id

            bulletin_bulletin = self.id

            period = dataQ[q][1]  

            record_value = { 'load_date': date_stat, 'bulletin_bulletin': bulletin_bulletin, 'industries': industries_id, 'period': period }
            self.env[ 'bulletin.load_date' ].create( record_value )  

            q = q + 1

# модель attachment вложения 
class bulletin_attachment(models.Model):

    _name = 'bulletin.attachment'
    _description = 'Присоединенные файлы'

    name = fields.Char( string = "Наименование" )
    attachment = fields.Binary( string = "Файл", attachment = False, required = True )
    bulletin = fields.Many2one( 'bulletin.bulletin', string = 'Статистика' )
    # date_load = fields.Date( string = "Дата загрузки", default = datetime.now().date() )     

    date_load = fields.Date( string = "Дата загрузки", default = lambda self: fields.datetime.now().date() ) 

# модель Статистика отбор statistics
class bulletin_statistics_select(models.Model):

    _name = 'bulletin.statistics_select'
    _description = 'Статистика отбор'
    _order = "date_stat"                # desc

    # текущий юзер
    res_users = fields.Many2one( 'res.users', string = 'user' ) 

    name = fields.Char( string = "Статистика" )

    date_stat = fields.Date( string = "Дата" ) 

    count = fields.Float( digits=( 10, 3 ), string = "Количество" )

    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатели отрасли' )
    characteristics_industries = fields.Many2one( 'bulletin.characteristics_industries', string = 'Характеристики отрасли' )

    region = fields.Many2one( 'bulletin.regions', string = 'Регион' )  
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона' ) 

    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2' )
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )

    quarter = fields.Boolean( string = 'Квартал' )

    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )

# прямой connect к базе result            
def ConnectToBase( loval_server ):

    if loval_server == True:
        try:
            conn = psycopg2.connect("host='localhost' dbname='result' user='postgres' password='password' ")
            return conn
        except:
            raise ValidationError( "Connection error !" )
            return None
    
    if loval_server == False:
        try:
            conn = psycopg2.connect("host='localhost' dbname='result' user='postgres' password='password' ")
            return conn
        except:
            raise ValidationError( "Connection error !" )
            return None

# wizard установка отбора статистики
class bulletin_stat_wizard(models.TransientModel):

    _name = 'bulletin.stat_wizard'
    _description = 'Установка отбора'

    stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Тип статистики', default = 1 )
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Вид статистики' )                       
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатель отрасли' )      
    characteristics_industries = fields.Many2one( 'bulletin.characteristics_industries', string = 'Характеристики отрасли' ) 
    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2', domain = lambda self:self.get_okved_domain() )
    message = fields.Char( string = "Комментарий"  )

    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность' )     

    def get_okved_domain( self ):
        domain = [ ( "status", "=", True ) ]
        return domain
    
    loval_server = fields.Boolean( string = 'Локальный сервер', default = True )
    quarter = fields.Boolean( string = 'Квартал' )
    
    region = fields.Many2many( 'bulletin.regions', string = "Регион" ) 
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона', default = 2 )

    not_null_number = fields.Boolean( string = 'Только заполненые №№', default = False )

    @api.onchange( 'region_type' )
    def change_region_type( self ):

        domain = { 'region': [] }
        for rec in self:
            if rec.region_type:
                domain = {'region': [ ('region_type.id', '=', rec.region_type.id ) ]}

                if rec.region_type.id == 3:
                    rf = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', 95 ) ] )
                    self.region = rf
        return {'domain': domain}


    date_start_view = fields.Many2one( 'bulletin.load_date', string = 'Период начало' )         # , default = 1 
    date_start = fields.Date( string = 'Период начало' )

    date_finish_view = fields.Many2one( 'bulletin.load_date', string = 'Период окончание' )     # , default = 16
    date_finish = fields.Date( string = 'Период окончание' )
    # -- 4 поля даты 

    @api.onchange( 'date_start' )
    def change_date_start( self ):
        for rec in self:
            if rec.date_start:
                rec.date_finish = rec.date_start 

    @api.onchange( 'industries' )
    def change_industries( self ):

        domain_start = { 'date_start_view': [] }        
        for rec in self:
            if rec.industries:
                rec.message = rec.industries.name 
                domain_start = {'date_start_view': [ ('industries.id', '=', rec.industries.id ) ], 'date_finish_view': [ ('industries.id', '=', rec.industries.id ) ], 'parameter_type': [ ('industries.id', '=', rec.industries.id ) ], 'indicators_industries': [ ('industries.id', '=', rec.industries.id ) ] }
                
        return { 'domain': domain_start }
        # -- установить отбор

    # в базу result для datalens из bulletin_statistics_select
    def select_to_datalens( self ):

        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )
        
        if self.industries.datalens == False:
            raise ValidationError( 'У статистики не определена таблица Datalens' )

        current_user_id = self.env.uid   

        # connect к базе result
        conn = ConnectToBase( self.loval_server )
        cursor = conn.cursor()
        
        ComStr = "SELECT date_stat, count, region, industries, indicators_industries FROM bulletin_statistics_select WHERE res_users = " + str( current_user_id ) 
        self._cr.execute( ComStr )

        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        industries = self.industries.name

        q = 0
        while q < CountQ:          

            date_stat = dataQ[q][0]
            count = dataQ[q][1]

            region_link = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', int( dataQ[q][2] ) ) ] )
            region = region_link.reg_name
            region_fedstat = region_link.fedstat
            reg_number = region_link.reg_number
            federal = region_link.federal.name

            if federal == None:
                federal = '' 
            
            indicators_industries = self.env[ 'bulletin.indicators_industries' ].search( [ ( 'id', '=', int( dataQ[q][4] ) ) ] ).name  

            MassivData = [ date_stat, count, industries, region, indicators_industries, region_fedstat, reg_number, federal ]
            sql = """INSERT INTO price ( date_stat, count, industries, region, indicators_industries, region_fedstat, reg_number, federal ) VALUES ( %s, %s, %s, %s, %s, %s, %s, %s )"""
            cursor.execute( sql, MassivData )
            conn.commit()

            q = q + 1

        conn.close()

    oper_type = fields.Selection(  [('1', 'Вывести количество'), ('2', 'Удалить данные')], "Действие"  )

    # test result
    def service( self ):
    
        if self.industries.id == False:
                return show_message( self, 'Тип статистики не определен' )

        if self.oper_type == '1':

            industries_tmp = str( self.industries.id )
            ComStr = "SELECT id FROM bulletin_statistics WHERE industries = " + industries_tmp 
            self._cr.execute( ComStr )
            CountQ = self._cr.rowcount

            return show_message( self, CountQ )

        elif self.oper_type == '2':

            # найти записи industries_tmp и удалить
            industries_tmp = str( self.industries.id )
            
            ComStr = "SELECT id FROM bulletin_statistics WHERE industries = " + industries_tmp  
            self._cr.execute( ComStr )

            CountQ = self._cr.rowcount
            dataQ = self._cr.fetchall()
        
            q = 0
            while q < CountQ:
                rec = self.env[ 'bulletin.statistics' ].search( [ ( 'id', '=', dataQ[q][0] ) ] )
                rec.unlink()
                
                q = q + 1              
    
        else:
            return show_message( self, 'Действие не определено' )
        

    def test_base( self ):
        pass

    # отбор из модели bulletin.statistics и запись в bulletin.statistics_select
    def select_to_base( self ):
        
        # Отрасль, проверить что указана, д.б. всегда
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )

        # # найти записи current_user_id и удалить
        current_user_id = self.env.uid
        rec = self.env[ 'bulletin.statistics_select' ].search( [ ( 'res_users', '=', current_user_id ) ] )
        rec.unlink()

        # регионы из списка self.region в подстроку
        if len( self.region ) > 0:
            RegStr = ' and (' 

            i = 0
            while i < len( self.region ):
                RegStr = RegStr + ' region = ' + str( self.region[i].id ) 
                if len( self.region ) > 1 and i < len( self.region ) - 1:
                    RegStr = RegStr + ' or'
                i = i + 1   
            RegStr = RegStr + ' ) '

        ComStr = "SELECT date_stat, count, region, industries, parameter_type, okved, indicators_industries, quarter, period FROM bulletin_statistics WHERE industries = " 
        ComStr = ComStr + str( self.industries.id )

        if len( self.region ) > 0:
            ComStr = ComStr + RegStr 

        if self.parameter_type.id != False:
            ComStr = ComStr + " and parameter_type = " + str( self.parameter_type.id )        
        if self.okved.id != False:
            ComStr = ComStr + " and okved = " + str( self.okved.id ) 
        if self.date_start != False:
            ComStr = ComStr + " and date_stat >= '" + str( self.date_start ) + "'"
        if self.date_finish != False:
            ComStr = ComStr + " and date_stat <= '" + str( self.date_finish ) + "'"
            
        if self.indicators_industries.id != False:
            ComStr = ComStr + " and indicators_industries = " + str( self.indicators_industries.id ) 

        if self.period.id != False:  
            ComStr = ComStr + " and period = " + str( self.period.id ) 

        ComStr = ComStr + " ORDER BY date_stat"

        self._cr.execute( ComStr )

        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        q = 0
        while q < CountQ:

            date_stat = dataQ[q][0]
            count = dataQ[q][1]
            region = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', int( dataQ[q][2] ) ) ] )

            if self.not_null_number == True and region.reg_number == 0:
                q = q + 1
                continue

            industries = self.industries

            quarter = dataQ[q][7]

            period = dataQ[q][8]

            if self.industries.industries_parameter.id == 1:
                parameter_type = self.env[ 'bulletin.parameter_type' ].search( [ ( 'id', '=', int( dataQ[q][4] ) ) ] )

                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'parameter_type': parameter_type.id, 'quarter': quarter, 'period': period }  

            if self.industries.industries_parameter.id == 3:
                parameter_type = self.env[ 'bulletin.parameter_type' ].search( [ ( 'id', '=', int( dataQ[q][4] ) ) ] )
                indicators_industries = self.env[ 'bulletin.indicators_industries' ].search( [ ( 'id', '=', int( dataQ[q][6] ) ) ] )

                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'parameter_type': parameter_type.id, 'quarter': quarter, 'period': period, 'indicators_industries': indicators_industries.id }  
                                
            if self.industries.industries_parameter.id == 4:      
                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'quarter': quarter, 'period': period }  
                
            if self.industries.industries_parameter.id == 5:
                parameter_type = self.env[ 'bulletin.parameter_type' ].search( [ ( 'id', '=', int( dataQ[q][4] ) ) ] )
                okved = self.env[ 'bulletin.okved' ].search( [ ( 'id', '=', int( dataQ[q][5] ) ) ] )

                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'parameter_type': parameter_type.id, 'industries': industries.id, 'okved': okved.id, 'quarter': quarter, 'period': period }  

            if self.industries.industries_parameter.id == 6:
                okved = self.env[ 'bulletin.okved' ].search( [ ( 'id', '=', int( dataQ[q][5] ) ) ] )
                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                 'industries': industries.id, 'okved': okved.id, 'quarter': quarter, 'period': period }  
                              
            if industries.id == 3 or industries.id == 4:
                indicators_industries = self.env[ 'bulletin.indicators_industries' ].search( [ ( 'id', '=', int( dataQ[q][6] ) ) ] )

                if count == 0 and region.archive == True:
                    pass
                else:    
                    record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                    'industries': industries.id, 'indicators_industries': indicators_industries.id }  

            if industries.id == 15 or industries.id == 10:
                indicators_industries = self.env[ 'bulletin.indicators_industries' ].search( [ ( 'id', '=', int( dataQ[q][6] ) ) ] )

                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'indicators_industries': indicators_industries.id, 'quarter': quarter }                  

            if industries.id == 17 or industries.id == 24:
                parameter_type = self.env[ 'bulletin.parameter_type' ].search( [ ( 'id', '=', int( dataQ[q][4] ) ) ] )

                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'parameter_type': parameter_type.id, 'quarter': quarter }  
                
            if industries.id == 18 or industries.id == 19 or industries.id == 20 or industries.id == 21 or industries.id == 9:
                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id }  
                                
            if industries.id == 6 :                
                record_value = { 'res_users': current_user_id, 'date_stat': date_stat, 'count': count, 'region': region.id, 'region_type': region.region_type.id,
                                'industries': industries.id, 'quarter': quarter }  
                
            self.env[ 'bulletin.statistics_select' ].create( record_value ) 
            
            q = q + 1

# wizard загрузки ...
class bulletin_load_wizard(models.TransientModel):

    _name = 'bulletin.load_wizard'
    _description = 'wizard загрузки'

    date_stat = fields.Date( string = 'Дата'  )  
    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Вид статистики'  )       
    characteristics_industries = fields.Many2one( 'bulletin.characteristics_industries', string = 'Характеристики статистики' )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатели статистики' )
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )
    stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2' )

    text1 = fields.Char( string = "Дата/Год", default = '20YY-MM-01' ) 
    text2 = fields.Char( string = "Имя файла" )   
    text3 = fields.Char( string = "Имя листа", default = 'Данные' )  
    letter = fields.Char( string = "Колонка"  )  

    start_index = fields.Integer( string = "Начальная строка" )    
    end_index = fields.Integer( string = "Конечная строка" ) 

    quarter = fields.Boolean( string = 'Квартал', default = False )
    message = fields.Char( string = "Комментарий"  )
    attachment = fields.Many2one( 'bulletin.attachment', string = 'Файл' )
    industries_stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    packet = fields.Boolean( string = 'Пакетная загрузка' )        

    period = fields.Many2one( 'bulletin.statistics_period', string = 'Периодичность', default = 1 )

    @api.onchange( 'industries' )
    def change_industries( self ):
        # установить отбор
        domain_start = { 'date_start_view': [] }
        for rec in self:
            if rec.industries:

                rec.message = rec.industries.industries_parameter.name
                rec.industries_stat_type = rec.industries.industries_stat_type
                domain_start = { 'parameter_type': [ ('industries.id', '=', rec.industries.id ) ], 'indicators_industries': [ ('industries.id', '=', rec.industries.id ) ] }
                
                self._cr.execute( "SELECT MAX( id ) FROM bulletin_load_wizard WHERE industries = %d " %( rec.industries.id ) )
                dataMax = self._cr.fetchall()
                id_max = dataMax[0][0]
                CountMax = self._cr.rowcount

                if CountMax == 1:
                    max_record = self.env[ 'bulletin.load_wizard' ].search( [ ( 'id', '=', id_max ) ] )
                    rec.text2 = max_record.text2
                    rec.start_index = max_record.start_index
                    rec.period = max_record.period
                    rec.end_index = max_record.end_index

        return { 'domain': domain_start }

 
    @api.onchange( 'packet' )
    def change_packet( self ):
        for rec in self:
            if rec.packet:
                rec.text1 = ''    

    @api.onchange( 'parameter_type' )
    def change_parameter_type( self ):
        for rec in self:
            if rec.parameter_type:
                if rec.parameter_type.sheet_name != False:
                    rec.text3 = rec.parameter_type.sheet_name.strip()

    @api.onchange( 'indicators_industries' )
    def change_indicators_industries( self ):
        for rec in self:
            if rec.indicators_industries:
                if rec.indicators_industries.sheet_name != False:
                    rec.text3 = rec.indicators_industries.sheet_name.strip()

    
    def get_load_rosstat( self ):

        # Объем платных услуг населению
        if self.industries.id == 14:
            self.get_load_rosstat_14( self)
        # Оборот розничной торговли
        if self.industries.id == 13:
            self.get_load_rosstat_13( self)
        # Индексы потребительских цен
        if self.industries.id == 15:
            self.get_load_rosstat_15()
        # Сальдированный финансовый результат, Прибыль прибыльных организаций       ( parameter_type )  	Среднемесячная номинальная заработная плата работников
        if self.industries.id == 16 or self.industries.id == 17 or self.industries.id == 22:
            self.get_load_rosstat_16()
        # Доля прибыльных предприятий и организаций, Доля убыточных предприятий и организаций, Индексы цен производителей промышленных товаров, Индексы цен на грузовые перевозки
        if self.industries.id == 18 or self.industries.id == 19 or self.industries.id == 20 or self.industries.id == 21:
            self.get_load_rosstat_18()
        # Среднедушевые денежные доходы населения, Объем инвестиций в основной капитал, Грузооборот автомобильного транспорта, Ввод жилья - or self.industries.id == 23
        if self.industries.id == 8 or self.industries.id == 5 or self.industries.id == 12:
            self.get_load_rosstat_parameter_type_quarter()

    # parameter_type + quarter
    def create_record_parameter_type_quarter(self, year_stat, month, val, rec_region, industries_id, parameter_type_id, quarter ):
            date_stat = str(year_stat) + month
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

            try:
                float(val)
            except ValueError:
                val = 0

            record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'parameter_type': parameter_type_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }
            self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_parameter_type_quarter( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )         
        if self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' )         
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )      
        
        # вид статистики и тип параметра
        industries_id = self.industries.id
        parameter_type_id = self.parameter_type.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 

        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )

        wb = load_workbook( filestr )

        if self.text3 not in wb.sheetnames:
            raise ValidationError( 'Лист не найден' )
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = sheet.max_row

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион по полю для Росстата eng_name
            valA = str( sheet['A' + str( i )].value ).strip()
            valA = valA.replace("\r","")
            valA = valA.replace("\n","")
            
            # Росстат - eng_name
            rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'eng_name', '=', valA ) ] )                        

            if rezalt == 0:
                raise ValidationError( 'Не найден регион ' + valA + ' строка ' + str( i ) )            
            i = i + 1
        # \\ проверки

        # квартал / не квартал
        quarter = self.quarter

        # год начала ведения статистики 
        start_year = int( self.text1 )

        # количество колонок - полных лет * 4 квартала + кол-во кварталов последнего неполного года + 1 ( сдвиг первой колонки с регионом )
        if quarter == True:
            x = ( 2023 - start_year ) * 4 + int( self.letter ) + 1
        else:
            x = ( 2023 - start_year ) * 12 + int( self.letter ) + 1

        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]   
        fin_list = []

        i = self.start_index

        counter = 0
        year_stat = int( self.text1 )
        
        while i <= max_row:

            counter = 0
            year_stat = int( self.text1 )
            
            valA = str( sheet['A' + str( i )].value ).strip()
            valA = valA.replace("\r","")
            valA = valA.replace("\n","")

            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'eng_name', '=', valA ) ] )     # Росстат - eng_name

            for j in list_of_letters:
              
                if quarter == True:
                    if counter != 0:
                        counter += 3
                    else:
                        counter += 1

                    if counter == 13:
                        counter = 1
                        year_stat += 1

                    month = "-0" + str(counter) + "-01"

                    if counter == 10:
                        month = "-" + str(counter) + "-01"
                else:

                    counter += 1

                    if counter == 13:
                        counter = 1
                        year_stat += 1

                    month = "-0" + str(counter) + "-01"

                    if counter >= 10:
                        month = "-" + str(counter) + "-01"

                value = sheet[j + str( i )].value

                self.create_record_parameter_type_quarter( year_stat, month, str(value) , rec_region, industries_id, parameter_type_id, quarter )
            i+=1    
    # -- Сальдированный финансовый результат, , Прибыль прибыльных организаций

    def create_trade_record_18(self, year_stat, month, val, rec_region, industries_id ):
            date_stat = str(year_stat) + month
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

            try:
                float(val)
            except ValueError:
                val = 0

            record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id }
            self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_18( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )         
       
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )      
        
        industries_id = self.industries.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 

        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )

        wb = load_workbook( filestr )
        # читать лист
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = sheet.max_row

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
            if rezalt == 0:
                raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
            i = i + 1
        # \\ проверки

        # количество колонок - полных лет * 12 месяцев + кол-во месяцев последнего неполного года + 1 ( сдвиг первой колонки с регионом )
        start_year = int( self.text1 )
        x = ( 2023 - start_year ) * 12 + int( self.letter ) + 1

        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]          # удаление элемента по индексу или ключу в случае словаря
        fin_list = []

        i = self.start_index

        counter = 0
        year_stat = int( self.text1 )
        
        while i <= max_row:

            counter = 0
            year_stat = int( self.text1 )
            
            valA = str( sheet['A' + str( i )].value ).strip()
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            for j in list_of_letters:
                counter += 1
                if counter == 13:
                    counter = 1
                    year_stat += 1
                month = "-0" + str(counter) + "-01"
                if counter > 9:
                    month = "-" + str(counter) + "-01"
                value = sheet[j + str( i )].value

                self.create_trade_record_18( year_stat, month, str(value) , rec_region, industries_id )
            i+=1    
    # -- Доля прибыльных предприятий и организаций, Доля убыточных предприятий и организаций

    # Сальдированный финансовый результат, , Прибыль прибыльных организаций
    def create_trade_record_16(self, year_stat, month, val, rec_region, industries_id, parameter_type_id ):
            date_stat = str(year_stat) + month
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

            try:
                float(val)
            except ValueError:
                val = 0

            record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'parameter_type': parameter_type_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id }
            self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_16( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )         
        if self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' )         
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )      
        
        industries_id = self.industries.id
        parameter_type_id = self.parameter_type.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 

        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )

        wb = load_workbook( filestr )
        # читать лист
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = sheet.max_row

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
            if rezalt == 0:
                raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
            i = i + 1
        # \\ проверки

        # количество колонок - полных лет * 12 месяцев + кол-во месяцев последнего неполного года + 1 ( сдвиг первой колонки с регионом )
        start_year = int( self.text1 )
        x = ( 2023 - start_year ) * 12 + int( self.letter ) + 1

        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]          # удаление элемента по индексу или ключу в случае словаря
        fin_list = []

        i = self.start_index

        counter = 0
        year_stat = int( self.text1 )
        
        while i <= max_row:

            counter = 0
            year_stat = int( self.text1 )
            
            valA = str( sheet['A' + str( i )].value ).strip()
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            for j in list_of_letters:
                counter += 1
                if counter == 13:
                    counter = 1
                    year_stat += 1
                month = "-0" + str(counter) + "-01"
                if counter > 9:
                    month = "-" + str(counter) + "-01"
                value = sheet[j + str( i )].value

                self.create_trade_record_16( year_stat, month, str(value) , rec_region, industries_id, parameter_type_id )
            i+=1    
    # -- Сальдированный финансовый результат, , Прибыль прибыльных организаций

    # Индексы потребительских цен
    def create_trade_record_15(self, year_stat, month, val, rec_region, industries_id, indicators_industries ):
            date_stat = str(year_stat) + month
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

            if val == None:
                val = 0
            else:
                try:
                    float(val)
                except ValueError:
                    val = 0

            record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'indicators_industries': indicators_industries, 'region': rec_region.id, 'region_type': rec_region.region_type.id }
            self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_15( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )         
        if self.indicators_industries.id == False:
            raise ValidationError( 'Не указан Показатель' )         
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )      
        
        industries_id = self.industries.id
        indicators_industries = self.indicators_industries.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 

        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )

        wb = load_workbook( filestr )
        # читать лист
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = 100

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
            if rezalt == 0:
                raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
            i = i + 1
        # \\ проверки

        # количество колонок - полных лет * 12 месяцев + кол-во месяцев последнего неполного года + 1 ( сдвиг первой колонки с регионом )
        start_year = int( self.text1 )
        x = ( 2023 - start_year ) * 12 + int( self.letter ) + 1

        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]          # удаление элемента по индексу или ключу в случае словаря
        fin_list = []

        i = self.start_index

        counter = 0
        year_stat = int( self.text1 )
        
        while i <= max_row:

            counter = 0
            year_stat = int( self.text1 )
            
            valA = str( sheet['A' + str( i )].value ).strip()
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            for j in list_of_letters:
                counter += 1
                if counter == 13:
                    counter = 1
                    year_stat += 1
                month = "-0" + str(counter) + "-01"
                if counter > 9:
                    month = "-" + str(counter) + "-01"
                value = sheet[j + str( i )].value

                self.create_trade_record_15( year_stat, month, str(value) , rec_region, industries_id, indicators_industries )
            i+=1    
    # -- Индексы потребительских цен

    # Объем платных услуг населению
    def create_trade_record_14(self, year_stat, month, val, rec_region, industries_id, parameter_type_id ):
        date_stat = str(year_stat) + month
        date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

        if val == None:
            val = 0
        else:
            try:
                float(val)
            except ValueError:
                val = 0

        record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'parameter_type': parameter_type_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id }
        self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_14( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  

        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' ) 
        if self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' ) 
        
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )
        
        industries_id = self.industries.id
        parameter_type_id = self.parameter_type.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 

        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )

        wb = load_workbook( filestr )
        # читать лист
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = sheet.max_row    

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
            if rezalt == 0:
                raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
            i = i + 1
        # \\ проверки

        x = 79
        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]          # удаление элемента по индексу или ключу в случае словаря
        fin_list = []

        # i = 5
        i = self.start_index

        counter = 0
        year_stat = int( self.text1 )
        
        while i <= max_row:

            counter = 0
            year_stat = int( self.text1 )
            
            valA = str( sheet['A' + str( i )].value ).strip()
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            for j in list_of_letters:
                counter += 1
                if counter == 13:
                    counter = 1
                    year_stat += 1
                month = "-0" + str(counter) + "-01"
                if counter > 9:
                    month = "-" + str(counter) + "-01"
                value = sheet[j + str( i )].value

                self.create_trade_record_14(year_stat, month, str(value) , rec_region, industries_id, parameter_type_id )
            i+=1    
    # -- Объем платных услуг населению

    # Оборот розничной торговли
    def create_trade_record_13(self, year_stat, month, val, rec_region, industries_id, parameter_type_id, trade_type):
        date_stat = str(year_stat) + month
        date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

        if val == None:
            val = 0
        else:
            try:
                float(val)
            except ValueError:
                val = 0

        record_value = { 'count': val, 'date_stat': date_stat, 'industries': industries_id, 'parameter_type': parameter_type_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'indicators_industries': trade_type}
        self.env[ 'bulletin.statistics' ].create( record_value )

    def get_load_rosstat_13( self ):

        industries_id = self.industries.id
        parameter_type_id = self.parameter_type.id
        trade_type = self.indicators_industries.id

        # читать книгу
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
        wb = load_workbook( filestr )
        # читать лист
        sheet = wb[self.text3]

        # кол-во строк на листе
        max_row = sheet.max_row    

        x = 79
        list_of_letters = [string.ascii_uppercase[i] if i < 26 else string.ascii_uppercase[i // 26 - 1] + string.ascii_uppercase[i % 26] for i in range(x)]
        del list_of_letters[0]          # удаление элемента по индексу или ключу в случае словаря
        fin_list = []

        i = 5
        counter = 0
        year_stat = 2017
        
        while i <= 100:

            counter = 0
            year_stat = 2017
            
            valA = str( sheet['A' + str( i )].value ).strip()
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            for j in list_of_letters:
                counter += 1
                if counter == 13:
                    counter = 1
                    year_stat += 1
                month = "-0" + str(counter) + "-01"
                if counter > 9:
                    month = "-" + str(counter) + "-01"
                value = sheet[j + str( i )].value

                if i == 5 and self.text3.strip() == 'Млн. рублей':
                    value = round(value * 1000 , 2)

                self.create_trade_record_13(year_stat, month, str(value) , rec_region, industries_id, parameter_type_id, trade_type)
            i+=1    
    # -- Оборот розничной торговли

    # -- rosstat

    def get_month_stat( self ):

        if self.text1 == '20YY-MM-01':
            raise ValidationError( 'Не указана дата' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )                  
        if self.start_index == 0:
            raise ValidationError( 'Не указана Начальная строка' )              
        if str( self.letter ) == 'False' and self.packet == False:
            raise ValidationError( 'Не указана Колонка' )  
        
        if self.period.id == False:
            raise ValidationError( 'Не указана Периодичность статистики' ) 

        # если у статистики свойство - Тип параметра, то проверять что оно заполнено
        if self.industries.industries_parameter.id == 1 and self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' )       
        # если у статистики свойство - Показатели статистики, то проверять что оно заполнено 
        if self.industries.industries_parameter.id == 2 and self.indicators_industries.id == False:
            raise ValidationError( 'Не указан Показатель статистики' )    
          
        # если у статистики свойство 3 - Тип параметра + Показатели отрасли
        if self.industries.industries_parameter.id == 3 and self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' )       
        if self.industries.industries_parameter.id == 3 and self.indicators_industries.id == False:
            raise ValidationError( 'Не указан Показатель статистики' )   
        
        if self.industries.industries_parameter.id == 2 and self.indicators_industries.id == False:
            raise ValidationError( 'Не указан Показатель статистики' )   
        
        # если у статистики свойство - Тип параметра + ОКВЭД, то проверять что Тип параметра заполнено
        if self.industries.industries_parameter.id == 5 and self.parameter_type.id == False:
            raise ValidationError( 'Не указан Тип параметра' )       
                
        # источник данных, автоматом из статистики, можно изменить вручную - 'rosstat.gov.ru' / 'fedstat.ru' / ...
        industries_stat_type = self.industries_stat_type

        # наличие файла 
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
        FileExists = os.path.exists( filestr )
        if FileExists == False:
            raise ValidationError( 'Не найден файл ' + filestr )
        
        # читать книгу
        wb = load_workbook( filestr )
        # читать лист
        if self.text3 not in wb.sheetnames:
            raise ValidationError( 'Лист не найден' )
        sheet = wb[self.text3]

        max_row = self.end_index 

        # тип статистики
        industries_id = self.industries.id

        # если у статистики свойство - Тип параметра или Тип параметра + ОКВЭД, присвоить тип параметра
        if self.industries.industries_parameter.id == 1 or self.industries.industries_parameter.id == 5:  
            parameter_type_id = self.parameter_type.id            

        # если у статистики свойство - Тип параметра + Показатель отрасли, присвоить тип параметра и Показатель отрасли
        if self.industries.industries_parameter.id == 3:  
            parameter_type_id = self.parameter_type.id      
            indicators_industries_id = self.indicators_industries.id

        # если у статистики свойство - Показатель отрасли, присвоить Показатель отрасли
        if self.industries.industries_parameter.id == 2:    
            indicators_industries_id = self.indicators_industries.id

        # квартал
        quarter = self.quarter

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            valA = valA.replace("\r","")
            valA = valA.replace("\n","")

            # ОКВЭД valA может быть None
            if self.industries.industries_parameter.id == 5 or self.industries.industries_parameter.id == 6:
                if valA != 'None':
                    rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                    if rezalt == 0:
                        rezalt = self.env[ 'bulletin.regions_name' ].search_count( [ ( 'name', '=', valA ) ] )
                        if rezalt == 0:
                            raise ValidationError( 'Не найден регион *' + str( valA ) + '* строка ' + str( i ) )
            else:
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    rezalt = self.env[ 'bulletin.regions_name' ].search_count( [ ( 'name', '=', valA ) ] )
                    if rezalt == 0:

                        x = valA
                        result = re.sub(r'[^А-я0-9-(): ]', '', x)
                        result2 = re.sub(" +", " ", result)
                        valA = result2

                        rezalt = self.env[ 'bulletin.regions_name' ].search_count( [ ( 'name', '=', valA ) ] )
                        if rezalt == 0:
                            raise ValidationError( 'Не найден регион *' + str( valA ) + '* строка ' + str( i ) )                            

            i = i + 1
        # \\ проверки

        # если пакетная загрузка
            
        if self.packet:

            if self.letter == None:
                raise ValidationError( 'Для пакетной загрузки нужно указать последнюю + 1 колонку' ) 

            record_value = ''

            if self.period.id == 1:

                letter_list = [ 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 
                                'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 
                                'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
                                'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW',
                                'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI',
                                'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU',
                                'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG',
                                'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS',
                                'CT'
                                ]
                              
            if self.period.id == 2:        
                letter_list = [ 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 
                                'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 
                                'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG'
                                ]                                   
                
            index_column = 0
            index_month = 1

            index_year = int( self.text1.strip() )

            # цикл по колонкам
            while index_column < len( letter_list ):

                if letter_list[index_column] == self.letter: 
                    break                    

                if index_month < 10:
                    date_stat = str( index_year ) + '-0' + str( index_month ) + '-01'
                else:
                    date_stat = str( index_year ) + '-' + str( index_month ) + '-01'

                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

                # + 1 если месяц, сбрасывать в начале след года
                if self.period.id == 1:

                    index_month = index_month + 1
                    if index_month == 13:
                        index_month = 1
                        index_year = index_year + 1

                # + 3 если квартал
                if self.period.id == 2:
                    index_month = index_month + 3

                    if index_month > 12:
                        index_month = 1
                        index_year = index_year + 1

                # цикл по строкам
                tmpValA = None

                i = self.start_index
                while i <= max_row:

                    # Регион
                    valA = str( sheet['A' + str( i )].value ).strip()
                    valA = valA.replace("\r","")
                    valA = valA.replace("\n","")

                    if self.industries.industries_parameter.id == 5 or self.industries.industries_parameter.id == 6:
                        if valA == 'None':
                            valA = tmpValA
                        else:
                            tmpValA = valA
                        rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )
                    else:
                        # регион из regions
                        rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                        # если не найдено в regions, регион из regions_name
                        if rec_region.id == False:
                            rec_region = self.env[ 'bulletin.regions_name' ].search( [ ( 'name', '=', valA ) ] )  
                            
                            if rec_region.id == False: 

                                x = valA
                                result = re.sub(r'[^А-я0-9-(): ]', '', x)
                                result2 = re.sub(" +", " ", result)
                                valA = result2

                                rec_region = self.env[ 'bulletin.regions_name' ].search( [ ( 'name', '=', valA ) ] )

                                if rec_region.id == False:                                
                                    raise ValidationError( 'Не найден регион *' + str( valA ) + '* строка ' + str( i ) ) 
                                
                                rec_region = rec_region.region
                            
                            else: 
                                rec_region = rec_region.region            
                        else:
                            rec_region = rec_region

                        # колонка с данными о количестве                        
                        valC = sheet[letter_list[index_column] + str( i )].value
                        if valC == None:
                            valC = 0 
                        try:
                            float( valC )
                        except ValueError:
                            valC = 0

                        # индивидуально
                            
                        # Оборот оптовой торговли, Оборот общественного питания - РФ * 1000 - коэф. в parameter_type.rf
                        if rec_region.id == 95 and self.parameter_type.rf > 0:
                            valC = valC * self.parameter_type.rf

                        # Тип параметра 
                        if self.industries.industries_parameter.id == 1:
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id  }  

                        # Показатель отрасли
                        if self.industries.industries_parameter.id == 2:
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period' : self.period.id, 'indicators_industries' : indicators_industries_id  } 

                        # Тип параметра + Показатель отрасли
                        if self.industries.industries_parameter.id == 3:
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id, 'indicators_industries' : indicators_industries_id  } 

                        # Неопределено
                        if self.industries.industries_parameter.id == 4:
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period' : self.period.id  } 

                        # Тип параметра + ОКВЭД
                        if self.industries.industries_parameter.id == 5:
                            valB = str( sheet['B' + str( i )].value ).strip()
                            okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
                            if okved.id == False:
                                raise ValidationError( 'Не найден ОКВЭД *' + str( valB ) + '* строка ' + str( i ) )
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id  }  

                        # Неопределено + ОКВЭД
                        if self.industries.industries_parameter.id == 6:
                            valB = str( sheet['B' + str( i )].value ).strip()
                            okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
                            if okved.id == False:
                                raise ValidationError( 'Не найден ОКВЭД *' + str( valB ) + '* строка ' + str( i ) )
                            record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period': self.period.id  }  

                        self.env[ 'bulletin.statistics' ].create( record_value ) 

                    i = i + 1
                # -- цикл по строкам             
                            
                index_column = index_column + 1

        else:

            tmpValA = None

            i = self.start_index
            while i <= max_row:

                # Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                valA = valA.replace("\r","")
                valA = valA.replace("\n","")

                # регион 
                # если ОКВЭД и valA может быть 'None'
                if self.industries.industries_parameter.id == 5 or self.industries.industries_parameter.id == 6:
                    if valA == 'None':
                        valA = tmpValA
                    else:
                        tmpValA = valA
                    rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )
                else:
                    # регион из regions
                    rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                    # если не найдено в regions, регион из regions_name
                    if rec_region.id == False:
                        rec_region = self.env[ 'bulletin.regions_name' ].search( [ ( 'name', '=', valA ) ] )  
                        
                        if rec_region.id == False: 

                            x = valA
                            result = re.sub(r'[^А-я0-9-(): ]', '', x)
                            result2 = re.sub(" +", " ", result)
                            valA = result2

                            rec_region = self.env[ 'bulletin.regions_name' ].search( [ ( 'name', '=', valA ) ] )

                            if rec_region.id == False:                                
                                raise ValidationError( 'Не найден регион *' + str( valA ) + '* строка ' + str( i ) ) 
                            
                            rec_region = rec_region.region
                        
                        else: 
                            rec_region = rec_region.region            
                    else:
                        rec_region = rec_region
                    
                # колонка с данными о количестве
                valC = sheet[self.letter.strip() + str( i )].value
                if valC == None:
                    valC = 0 
                try:
                    float( valC )
                except ValueError:
                    valC = 0
                
                # дата одна из поля text1
                date_stat = self.text1.strip()
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

                if rec_region.id == 95 and self.parameter_type.rf > 0:
                    valC = valC * self.parameter_type.rf

                # Тип параметра 
                if self.industries.industries_parameter.id == 1:
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id  }  

                # Показатель отрасли
                if self.industries.industries_parameter.id == 2:
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period' : self.period.id, 'indicators_industries' : indicators_industries_id  } 

                # Тип параметра + Показатель отрасли
                if self.industries.industries_parameter.id == 3:
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id, 'indicators_industries' : indicators_industries_id  } 

                # Неопределено
                if self.industries.industries_parameter.id == 4:
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period' : self.period.id  } 

                # Тип параметра + ОКВЭД
                if self.industries.industries_parameter.id == 5:
                    valB = str( sheet['B' + str( i )].value ).strip()
                    okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
                    if okved.id == False:
                        raise ValidationError( 'Не найден ОКВЭД *' + str( valB ) + '* строка ' + str( i ) )
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id, 'period' : self.period.id  }  

                # Неопределено + ОКВЭД
                if self.industries.industries_parameter.id == 6:
                    valB = str( sheet['B' + str( i )].value ).strip()
                    okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
                    if okved.id == False:
                        raise ValidationError( 'Не найден ОКВЭД *' + str( valB ) + '* строка ' + str( i ) )
                    record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'period': self.period.id  }  

                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1

    # -- ежемесячная загрузка

    # загрузить Fedstat
    def get_load_stat( self ):

        if self.industries.id == False:
            raise ValidationError( 'Не указан Вид статистики' )
        
        if self.industries.id == 12:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )         
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
            if self.letter == False:
                raise ValidationError( 'Не указано колонка с данными о количестве' )
            if self.start_index == 0:
                raise ValidationError( 'Не указана Начальная строка' )
                   
            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )

            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   

            # тип статистики
            industries_id = self.industries.id
            # квартал
            quarter = self.quarter

            parameter_type_id = self.parameter_type.id 

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1
            # \\ проверки

            i = self.start_index
            while i <= max_row:

                valA = str( sheet['A' + str( i )].value ).strip()   # Регион

                # Регион
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                # колонка с данными о количестве
                valC = sheet[self.letter.strip() + str( i )].value

                if quarter == True:

                    if self.letter.strip() == 'B':
                        date_stat = self.text1.strip() + '-01-01'
                    if self.letter.strip() == 'C':
                        date_stat = self.text1.strip() + '-04-01'
                    if self.letter.strip() == 'D':
                        date_stat = self.text1.strip() + '-07-01'
                    if self.letter.strip() == 'E':
                        date_stat = self.text1.strip() + '-10-01'
                else:

                    if self.letter.strip() == 'B':
                        date_stat = self.text1.strip() + '-01-01'
                    if self.letter.strip() == 'C':
                        date_stat = self.text1.strip() + '-02-01'
                    if self.letter.strip() == 'D':
                        date_stat = self.text1.strip() + '-03-01'
                    if self.letter.strip() == 'E':
                        date_stat = self.text1.strip() + '-04-01'
                    if self.letter.strip() == 'F':
                        date_stat = self.text1.strip() + '-05-01'
                    if self.letter.strip() == 'G':
                        date_stat = self.text1.strip() + '-06-01'
                    if self.letter.strip() == 'H':
                        date_stat = self.text1.strip() + '-07-01'                   

                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'parameter_type': parameter_type_id  }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1
        
        # -- Грузооборот автомобильного транспорта
        
        # Численность постоянного населения в среднем за год
        if self.industries.id == 3 or self.industries.id == 4:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )  
        
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
        
            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )

            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   

            # тип статистики
            industries_id = self.industries.id
            # показатель статистики
            indicators_industries = self.indicators_industries.id

            if indicators_industries == False:
                raise ValidationError( 'Не указан показатель статистики' )
            if self.start_index == 0:
                raise ValidationError( 'Не указана Начальная строка' )

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1
            # \\ проверки

            i = self.start_index
            while i <= max_row:

                valA = str( sheet['A' + str( i )].value ).strip()   # Регион

                # Регион
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                valC = sheet[self.letter.strip() + str( i )].value

                date_stat = self.text1.strip() + '-01-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'indicators_industries': indicators_industries  }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1

        # Численность занятых 15 лет и старше
        if self.industries.id == 6:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )        
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
        

            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )

            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            max_row = sheet.max_row   
            industries_id = self.industries.id
            quarter = self.quarter
        
            if self.start_index == 0:
                raise ValidationError( 'Не указана Начальная строка' )

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1
            # \\ проверки


            i = self.start_index
            while i <= max_row:

                valA = str( sheet['A' + str( i )].value ).strip()   # Регион

                # Регион
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                # 4 квартала + значение за год
                valC = sheet[self.letter.strip() + str( i )].value

                if quarter:

                    if self.letter.strip() == 'B':
                        date_stat = self.text1.strip() + '-03-01'
                    if self.letter.strip() == 'C':
                        date_stat = self.text1.strip() + '-06-01'
                    if self.letter.strip() == 'D':
                        date_stat = self.text1.strip() + '-09-01'
                    if self.letter.strip() == 'E':
                        date_stat = self.text1.strip() + '-12-01'
                else:

                    date_stat = self.text1.strip() + '-01-01'                 

                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()

                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1
           
        # Отгрузка товаров, услуг  https://fedstat.ru/indicator/57722
        if self.industries.id == 7:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )  
        
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
        
            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )

            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   

            # тип статистики
            industries_id = self.industries.id
            # тип параметра
            parameter_type_id = self.parameter_type.id            
            # квартал
            quarter = self.quarter
        
            if self.start_index == 0:
                raise ValidationError( 'Не указана Начальная строка' )

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                if valA != 'None':
                    rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                    if rezalt == 0:
                        raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                # проверить загружаемый файл статистики на ОКВЭД    
                valB = str( sheet['B' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.okved' ].search_count( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден ОКВЭД ' + str( valB ) + ' строка ' + str( i ) )
                i = i + 1

            # \\ проверки

            # цикл перебора строк, читать колонки

            tmpValA = None

            year_stat = self.text1.strip()

            i = self.start_index
            while i <= max_row:
                
                # Регион
                valA = str( sheet['A' + str( i )].value ).strip()  
                     
                if valA == 'None':
                    valA = tmpValA
                else:
                    rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] ) 
                    tmpValA = valA            

                # ОКВЭД
                valB = str( sheet['B' + str( i )].value ).strip() 
                okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )

                valC = sheet['C' + str( i )].value 
                valD = sheet['D' + str( i )].value 
                valE = sheet['E' + str( i )].value 
                valF = sheet['F' + str( i )].value 
                valG = sheet['G' + str( i )].value 
                valH = sheet['H' + str( i )].value 
                valI = sheet['I' + str( i )].value 

                if valC == None:
                    valC = 0.000
                if valD == None:
                    valD = 0.000
                if valE == None:
                    valE = 0.000
                if valF == None:
                    valF = 0.000
                if valG == None:
                    valG = 0.000
                if valH == None:
                    valH = 0.000
                if valI == None:
                    valI = 0.000

                # январь
                date_stat = year_stat + '-01-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # февраль
                date_stat = year_stat + '-02-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valD, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # март
                date_stat = year_stat + '-03-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valE, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # апрель
                date_stat = year_stat + '-04-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valF, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # май
                date_stat = year_stat + '-05-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valG, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июнь
                date_stat = year_stat + '-06-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valH, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июль
                date_stat = year_stat + '-07-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valI, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1

            # -- отгрузка

        # Реальная заработная плата работников организаций
        if self.industries.id == 9:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )  
        
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
        
            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )

            wb = load_workbook( filestr )

            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   

            # тип статистики
            industries_id = self.industries.id

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1
            # \\ проверки

            year_stat = self.text1.strip()

            i = self.start_index
            while i <= max_row:
                
                # Регион
                valA = str( sheet['A' + str( i )].value ).strip()  
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )                     

                valB = sheet['B' + str( i )].value  # январь-март 03-01
                valC = sheet['C' + str( i )].value  # январь-июнь 06-01

                if valB == None:
                    valB = 0.000
                if valC == None:
                    valC = 0.000

                # январь-март
                date_stat = year_stat + '-03-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valB, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # январь-июнь
                date_stat = year_stat + '-06-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 


                i = i + 1      
        # -- Реальная заработная плата работников организаций 

        # Индексы производства продукции сельского хозяйства
        if self.industries.id == 10:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )          
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
            if self.start_index == False:
                raise ValidationError( 'Не указана начальная строка' )
            if self.indicators_industries.id == False:
                raise ValidationError( 'Не указан показатель отрасли' )
            
            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )
            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   
            # тип статистики
            industries_id = self.industries.id
            # квартал
            quarter = self.quarter
            # показатель статистики
            indicators_industries = self.indicators_industries.id

            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1
            # \\ проверки

            year_stat = self.text1.strip()

            i = self.start_index
            while i <= max_row:
                
                # Регион
                valA = str( sheet['A' + str( i )].value ).strip()  
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )                     

                valB = sheet['B' + str( i )].value  # I квартал
                valC = sheet['C' + str( i )].value  # II квартал
                valD = sheet['D' + str( i )].value  # III квартал
                valE = sheet['E' + str( i )].value  # IV квартал

                if valB == None:
                    valB = 0.000
                if valC == None:
                    valC = 0.000
                if valD == None:
                    valD = 0.000
                if valE == None:
                    valE = 0.000

                # I квартал
                date_stat = year_stat + '-03-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valB, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'indicators_industries': indicators_industries }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # II квартал
                date_stat = year_stat + '-06-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'indicators_industries': indicators_industries  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # III квартал
                date_stat = year_stat + '-09-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valD, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'indicators_industries': indicators_industries  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # IV квартал
                date_stat = year_stat + '-12-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valE, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter, 'indicators_industries': indicators_industries  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                i = i + 1      
        # -- Индексы производства продукции сельского хозяйства

        # Объем работ, выполненных по виду деятельности Строительство    
        if self.industries.id == 11:

            if self.text1 == False:
                raise ValidationError( 'Не указан год' )         
            if self.text2 == False:
                raise ValidationError( 'Не указано имя файла' )
            if self.start_index == False:
                raise ValidationError( 'Не указана начальная строка' )

            # наличие файла 
            filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
            FileExists = os.path.exists( filestr )
            if FileExists == False:
                raise ValidationError( 'Не найден файл ' + filestr )
            
            # читать книгу
            wb = load_workbook( filestr )
            # читать лист
            if 'Данные' not in wb.sheetnames:
                raise ValidationError( 'Лист не найден' )
            sheet = wb['Данные']

            # кол-во строк на листе
            max_row = sheet.max_row   

            # тип статистики
            industries_id = self.industries.id        
            # квартал
            quarter = self.quarter
        
            # проверки
            i = self.start_index
            while i <= max_row:
                # проверить загружаемый файл статистики на Регион
                valA = str( sheet['A' + str( i )].value ).strip()
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) + ' строка ' + str( i ) )
                i = i + 1

            year_stat = self.text1.strip()

            i = self.start_index
            while i <= max_row:
                
                # Регион
                valA = str( sheet['A' + str( i )].value ).strip()  
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )                            

                valC = sheet['B' + str( i )].value  # январь
                valD = sheet['C' + str( i )].value  # февраль
                valE = sheet['D' + str( i )].value  # март
                valF = sheet['E' + str( i )].value  # апрель
                valG = sheet['F' + str( i )].value  # май
                valH = sheet['G' + str( i )].value  # июнь
                valI = sheet['H' + str( i )].value  # июль
                valJ = sheet['I' + str( i )].value  # август

                if valC == None:
                    valC = 0.000
                if valD == None:
                    valD = 0.000
                if valE == None:
                    valE = 0.000
                if valF == None:
                    valF = 0.000
                if valG == None:
                    valG = 0.000
                if valH == None:
                    valH = 0.000
                if valI == None:
                    valI = 0.000


                # январь
                date_stat = year_stat + '-01-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # февраль
                date_stat = year_stat + '-02-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valD, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # март
                date_stat = year_stat + '-03-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valE, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # апрель
                date_stat = year_stat + '-04-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valF, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # май
                date_stat = year_stat + '-05-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valG, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июнь
                date_stat = year_stat + '-06-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valH, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июль
                date_stat = year_stat + '-07-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valI, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # август
                date_stat = year_stat + '-08-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valJ, 'date_stat': date_stat, 'industries': industries_id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'quarter': quarter  }
                self.env[ 'bulletin.statistics' ].create( record_value ) 


                i = i + 1            
            # Объем работ, выполненных по виду деятельности Строительство        

    # загрузить ИПП в отдельную модель bulletin.ipp отдельной кнопкой get_load_ipp
    def get_load_ipp( self ):

        if self.text1 == False:
            raise ValidationError( 'Не указан год' )        
        if self.text2 == False:
            raise ValidationError( 'Не указан файл' )
        if self.text3 == False:
            raise ValidationError( 'Не указан лист' )  
        if self.parameter_type.id == False:
            raise ValidationError( 'Не указан тип параметра' )
        
        filestr = '/opt/odoo14/odoo/addons/bulletin/files/' + self.text2.strip() 
    
        wb = load_workbook( filestr )
        sheet = wb[self.text3]
        max_row = sheet.max_row    

        # проверки
        i = self.start_index
        while i <= max_row:
            # проверить загружаемый файл статистики на Регион
            valA = str( sheet['A' + str( i )].value ).strip()
            if valA != 'None':
                rezalt = self.env[ 'bulletin.regions' ].search_count( [ ( 'fedstat', '=', valA ) ] )
                if rezalt == 0:
                    raise ValidationError( 'Не найден регион ' + str( valA ) )
            # проверить загружаемый файл статистики на ОКВЭД
            valB = str( sheet['B' + str( i )].value ).strip()
            rezalt = self.env[ 'bulletin.okved' ].search_count( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )
            if rezalt == 0:
                raise ValidationError( 'Не найден ОКВЭД ' + str( valB ) )
            i = i + 1

        parameter_type_id = self.parameter_type.id
        year_stat = self.text1.strip()
        tmpValA = None
         
        i = self.start_index
        while i <= max_row:

            # Регион
            valA = str( sheet['A' + str( i )].value ).strip()   
            if valA == 'None':
                valA = tmpValA
            else:
                tmpValA = valA
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

            # ОКВЭД
            valB = str( sheet['B' + str( i )].value ).strip()   
            okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', valB ), ( 'status', '=', True ) ] )

            valC = sheet['C' + str( i )].value  # январь
            valD = sheet['D' + str( i )].value  # февраль
            valE = sheet['E' + str( i )].value  # март
            valF = sheet['F' + str( i )].value  # апрель
            valG = sheet['G' + str( i )].value  # май
            valH = sheet['H' + str( i )].value  # июнь
            valI = sheet['I' + str( i )].value  # июль

            valJ = sheet['J' + str( i )].value  # август
            valK = sheet['K' + str( i )].value  # сентябрь
            valL = sheet['L' + str( i )].value  # октябрь
            valM = sheet['M' + str( i )].value  # ноябрь
            valN = sheet['N' + str( i )].value  # декабрь

            if valC == None:
                valC = 0.000
            if valD == None:
                valD = 0.000
            if valE == None:
                valE = 0.000
            if valF == None:
                valF = 0.000
            if valG == None:
                valG = 0.000
            if valH == None:
                valH = 0.000
            if valI == None:
                valI = 0.000

            if valJ == None:
                valJ = 0.000
            if valK == None:
                valK = 0.000
            if valL == None:
                valL = 0.000
            if valM == None:
                valM = 0.000                    
            if valN == None:
                valN = 0.000

            # январь
            date_stat = year_stat + '-01-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valC, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id }  
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # февраль
            date_stat = year_stat + '-02-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valD, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # март
            date_stat = year_stat + '-03-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valE, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # апрель
            date_stat = year_stat + '-04-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valF, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # май
            date_stat = year_stat + '-05-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valG, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # июнь
            date_stat = year_stat + '-06-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valH, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # июль
            date_stat = year_stat + '-07-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valI, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # август
            date_stat = year_stat + '-08-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valJ, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # сентябрь
            date_stat = year_stat + '-09-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valK, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # октябрь
            date_stat = year_stat + '-10-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valL, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # ноябрь
            date_stat = year_stat + '-11-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valM, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            # декабрь
            date_stat = year_stat + '-12-01'
            date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
            record_value = { 'count': valN, 'date_stat': date_stat, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id  }
            self.env[ 'bulletin.ipp' ].create( record_value ) 

            i = i + 1

    # загрузить Сельское хозяйство из xlsx
    def get_load( self ):   
 
        wb = load_workbook( '/opt/odoo14/odoo/addons/bulletin/files/data.xlsx' )
        sheet = wb['Данные']
        max_row = sheet.max_row

        i = 5

        tmpValA = None
        tmpValB = None

        # цикл перебора строк, читать 4 колонки
        while i <= max_row:
            valA = sheet['A' + str( i )].value  
            valB = sheet['B' + str( i )].value   
            valC = sheet['C' + str( i )].value   
            valE = sheet['P' + str( i )].value   

            if valA != None:
                if tmpValA != valA:
                    tmpValA = str( valA ).strip()

            if valB != None:
                if tmpValB != valB:
                    tmpValB = str( valB ).strip()

            tmpValC = str( valC ).strip()

            if valE == None:
                tmpValE = 0.000
            elif valE == False:    
                tmpValE = 0.000
            elif valE == '':    
                tmpValE = 0.000
            else:
                tmpValE = float( valE )
            
            # Показатели отрасли 
            rec_ind_indust = self.env[ 'bulletin.indicators_industries' ].search( [ ( 'fedstat', '=', tmpValA ) ] )
            # Характеристики отрасли
            rec_charact_indust = self.env[ 'bulletin.characteristics_industries' ].search( [ ( 'fedstat', '=', tmpValB ) ] )
            # Регионы
            rec_regions = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', tmpValC ) ] )

            record_value = { 'date_stat': self.date_stat, 'count': tmpValE, 'industries': self.industries.id, 'indicators_industries': rec_ind_indust.id, 'characteristics_industries': rec_charact_indust.id, 'region': rec_regions.id, 'region_type': rec_regions.region_type.id  }

            self.env[ 'bulletin.statistics' ].create( record_value ) 

            i = i + 1

            x = 1 + 1

    # загрузить Агрегированный индекс производства
    def get_load_rosstart( self ):   

        wb = load_workbook( '/opt/odoo14/odoo/addons/bulletin/files/ipp-rosstat.xlsx' )
        # читать лист
        sheet = wb['3']      

        max_row = sheet.max_row    
        industries_id = self.industries.id
        parameter_type_id = self.parameter_type.id
        okved_id = self.okved.id

        stat_type_id = self.stat_type.id
        
        i = 6
        while i <= max_row:

            valA = str( sheet['A' + str( i )].value ).strip()   # Регион

            valC = sheet['BV' + str( i )].value  # январь
            valD = sheet['BW' + str( i )].value  # февраль
            valE = sheet['BX' + str( i )].value  # март
            valF = sheet['BY' + str( i )].value  # апрель
            valG = sheet['BZ' + str( i )].value  # май
            valH = sheet['CA' + str( i )].value  # июнь
            valI = sheet['CB' + str( i )].value  # июль
            valJ = sheet['CC' + str( i )].value  # август
            valK = sheet['CD' + str( i )].value  # сентябрь
            valL = sheet['CE' + str( i )].value  # октябрь
            valM = sheet['CF' + str( i )].value  # ноябрь
            valN = sheet['CG' + str( i )].value  # декабрь

            rezalt = self.env[ 'bulletin.okved' ].search_count( [ ( 'name', '=', 'Агрегированный индекс производства' ), ( 'status', '=', True ) ] )

            if rezalt == 0:
                raise ValidationError( 'Не найден Агрегированный индекс производства' ) 
            else:

                if valC == None:
                    valC = 0.000
                if valD == None:
                    valD = 0.000
                if valE == None:
                    valE = 0.000
                if valF == None:
                    valF = 0.000
                if valG == None:
                    valG = 0.000
                if valH == None:
                    valH = 0.000
                if valI == None:
                    valI = 0.000
                if valJ == None:
                    valJ = 0.000
                if valK == None:
                    valK = 0.000
                if valL == None:
                    valL = 0.000
                if valM == None:
                    valM = 0.000                    
                if valN == None:
                    valN = 0.000

                # ОКВЭД
                okved = self.env[ 'bulletin.okved' ].search( [ ( 'name', '=', 'Агрегированный индекс производства' ), ( 'status', '=', True ) ] )
                # Регион
                rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'fedstat', '=', valA ) ] )

                year_stat = '2021'

               # январь
                date_stat = year_stat + '-01-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valC, 'date_stat': date_stat, 'industries': industries_id, 'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }  
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # февраль
                date_stat = year_stat + '-02-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valD, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # март
                date_stat = year_stat + '-03-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valE, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # апрель
                date_stat = year_stat + '-04-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valF, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # май
                date_stat = year_stat + '-05-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valG, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июнь
                date_stat = year_stat + '-06-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valH, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # июль
                date_stat = year_stat + '-07-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valI, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # август
                date_stat = year_stat + '-08-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valJ, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # сентябрь
                date_stat = year_stat + '-09-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valK, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # октябрь
                date_stat = year_stat + '-10-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valL, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # ноябрь
                date_stat = year_stat + '-11-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valM, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

                # декабрь
                date_stat = year_stat + '-12-01'
                date_stat = datetime.strptime( date_stat, "%Y-%m-%d").date()
                record_value = { 'count': valN, 'date_stat': date_stat, 'industries': industries_id,  'okved': okved.id, 'region': rec_region.id, 'region_type': rec_region.region_type.id, 'parameter_type': parameter_type_id, 'stat_type': stat_type_id }
                self.env[ 'bulletin.statistics' ].create( record_value ) 

            i = i + 1


    def get_download_stat ( self ):

        rec = self.env[ 'bulletin.statistics_download' ].search( [ ( 'id', '>', 0) ] )
        rec.unlink()

        self._cr.execute( "SELECT date_stat, count, industries, okved, parameter_type, region FROM bulletin_statistics WHERE industries = 2 and date_stat > '31.12.2020' and date_stat < '01.01.2022' and parameter_type = 3 and ( region = 63 or region = 79 or region = 35 or region = 73 or region = 78 or region = 83 or region = 86 or region = 92 or region = 95 ) and ( okved = 181 or okved = 284 or okved = 1326 or okved = 1376 or okved = 2848  )  ORDER BY date_stat, region " )
        CountQ = self._cr.rowcount
        dataQ = self._cr.fetchall()

        industries = self.env[ 'bulletin.industries_economy' ].search( [ ( 'id', '=', 2 ) ] ) 

        q = 0
        while q < CountQ: 
                                   
            rec_okved = self.env[ 'bulletin.okved' ].search( [ ( 'id', '=', dataQ[q][3] ) ] )
            rec_region = self.env[ 'bulletin.regions' ].search( [ ( 'id', '=', dataQ[q][5] ) ] )

            record_value = { 'date_stat': str( dataQ[q][0] ), 'count': dataQ[q][1], 'industries': industries.id, 'parameter_type': 3, 'region': rec_region.id, 'okved': rec_okved.id, 'fedstat_region': rec_region.fedstat  }
            self.env[ 'bulletin.statistics_download' ].create( record_value )
            q = q + 1

# для вывода сообщения 
def show_message( self, message_text ):
        
        self.env['bulletin.message_view'].message_method( message_text ) 

        views = [  (self.env.ref('bulletin.form_message_view_action_window').id, 'form')  ]
        tree_view = {
            'name': 'Сообщение',
            'view_mode': 'form',
            'views': views,
            'res_model': 'bulletin.message_view',
            'type': 'ir.actions.act_window',   
            'target' : 'new',
        }
        return tree_view 

# модель message_view Просмотр сообщения 
class bulletin_message_view(models.TransientModel):

    tmp_list = list()

    _name = 'bulletin.message_view'
    _description = 'Просмотр сообщения'

    def message_method( self, message_text ):
        self.tmp_list.clear()
        self.tmp_list.append( message_text )            

    def get_basename( self ):
        return self.tmp_list[0]

    name = fields.Char( string = "Сообщение" )
    message = fields.Char( string = "Сообщение", default = get_basename  )

# модель diagram Диаграмма
class bulletin_diagram(models.Model):

    _name = 'bulletin.diagram'

    name = fields.Char( string = "name" ) 

    tutorial = fields.Char( string = "Заголовок" )

    # периоды
    Ox = fields.Char( string = "Ox" )
    # заголовки диаграмм - список
    Ot = fields.Char( string = "Ot" )
    # показатели диаграмм - списки
    Oy = fields.Char( string = "Oy" ) 

    table = fields.Many2many( 'bulletin.table', string = 'Таблица' ) 

    user = fields.Many2one( 'bulletin.workers', string = 'Пользователь' )

    date = fields.Date( string = 'Дата формирования' )

    # Сформировать выборку
    def write_rezalt(self):

        views = [  (self.env.ref('bulletin.form_stat_wizard_action_window').id, 'form')  ]
        tree_view = {
            'name': 'Статистика',
            'view_mode': 'form',
            'views': views,
            'res_model': 'bulletin.stat_wizard',
            'type': 'ir.actions.act_window',   
            'target' : 'new',
        }
        return tree_view 

# модель Статистика выгрузка statistics_download
class bulletin_statistics_download(models.Model):
    
    _name = 'bulletin.statistics_download'
    name = fields.Char( string = "Статистика" )

    # период
    date_stat = fields.Date( string = "Дата" ) 

    count = fields.Float( digits=( 10, 3 ), string = "Количество" )

    industries = fields.Many2one( 'bulletin.industries_economy', string = 'Тип статистики' )
    indicators_industries = fields.Many2one( 'bulletin.indicators_industries', string = 'Показатели отрасли' )
    characteristics_industries = fields.Many2one( 'bulletin.characteristics_industries', string = 'Характеристики отрасли' )
    region = fields.Many2one( 'bulletin.regions', string = 'Регион' )  
    fedstat_region = fields.Char( string = 'Регион' )
    region_type = fields.Many2one( 'bulletin.region_type', string = 'Тип региона' ) 
    okved = fields.Many2one( 'bulletin.okved', string = 'ОКВЭД 2' )
    parameter_type = fields.Many2one( 'bulletin.parameter_type', string = 'Тип параметра' )
    stat_type = fields.Many2one( 'bulletin.stat_type', string = 'Источник данных' )
    column = fields.Char( string = "Колонка" )













                         






